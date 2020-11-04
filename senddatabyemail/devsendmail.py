# encoding=utf-8
import traceback
import pymysql
import openpyxl
import datetime
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.header import Header
import smtplib


class mail:
    def __init__(self):
        pass
    def create_email(self,email_from, email_to, email_title, email_text, annex_path=None, annex_name=None):
        """
        生成一个空的带附件的邮件实例
        :param email_from: 发件人名称；可非邮箱
        :param email_to:   收件人名称；可非邮箱
        :param email_title: 邮件标题
        :param email_text:  邮件正文
        :param annex_path:  附件的本地目录
        :param annex_name:  定义附件名
        :return:
        """
        message = MIMEMultipart()
        #将正文以text的形式插入邮件中
        message.attach(MIMEText(email_text, 'html', 'utf-8'))
        #生成发件人名称（这个跟发送的邮件没有关系）
        message['From'] = Header(email_from, 'utf-8')
        #生成收件人名称（这个跟接收的邮件也没有关系）
        message['To'] = Header(email_to, 'utf-8')
        #生成邮件主题
        message['Subject'] = Header(email_title, 'utf-8')
        if annex_path is not None:
            #读取附件的内容
            with open(annex_path, 'rb') as file:
                content = file.read()
            att1 = MIMEText(content,'base64','utf-8')
            att1["Content-Type"] = 'application/octet-stream'
            #生成附件的名称
            att1["Content-Disposition"] = 'attachment; filename=' + annex_name
            #将附件内容插入邮件中
            message.attach(att1)
        #返回邮件
        return message

    def send_email(self,sender, password, receiver, msg):
        # 一个输入邮箱、密码、收件人、邮件内容发送邮件的函数
        try:
            #找到你的发送邮箱的服务器地址，已加密的形式发送
            server = smtplib.SMTP_SSL("smtp.exmail.qq.com", 465)  # 发件人邮箱中的SMTP服务器
            server.ehlo()
            #登录你的账号
            server.login(sender, password)  # 括号中对应的是发件人邮箱账号、邮箱密码
            #发送邮件
            server.sendmail(sender, receiver, msg.as_string())  # 括号中对应的是发件人邮箱账号、收件人邮箱账号（是一个列表）、邮件内容
            print("邮件发送成功")
            server.quit()  # 关闭连接
        except Exception:
            print(traceback.print_exc())
            print("邮件发送失败")



class datas:
    def __init__(self):
        self.m = mail()
    def get_datas(self,sql):
        # 一个传入sql导出数据的函数
        # 跟数据库建立连接
        conn = pymysql.connect(host='123.56.126.125', user='txy',
                              passwd='123456', database='zentao', port=3306, charset="utf8")
        # 使用 cursor() 方法创建一个游标对象 cursor
        cur = conn.cursor()
        # 使用 execute() 方法执行 SQL
        cur.execute(sql)
        # 获取所需要的数据
        datas = cur.fetchall()
        #关闭连接
        cur.close()
        #返回所需的数据
        return datas

    def get_fields(self,sql):
        # 一个传入sql导出字段的函数
        conn = pymysql.connect(host='123.56.126.125', user='txy',
                              passwd='123456', database='zentao', port=3306, charset="utf8")
        cur = conn.cursor()
        cur.execute(sql)
        # 获取所需要的字段名称
        fields = cur.description
        cur.close()
        return fields

    def get_excel(self,data, field, file):
        # 将数据和字段名写入excel的函数
        #新建一个工作薄对象
        new = openpyxl.Workbook()
        #激活一个新的sheet
        sheet = new.active
        #给sheet命名
        sheet.title = 'bug数据展示'
        #将字段名称循环写入excel第一行，因为字段格式列表里包含列表，每个列表的第一元素才是字段名称
        for col in range(len(field)):
            #row代表行数，column代表列数，value代表单元格输入的值，行数和列数都是从1开始，这点于python不同要注意
            _ = sheet.cell(row=1, column=col+1, value=u'%s' % field[col][0])
         #将数据循环写入excel的每个单元格中
        for row in range(len(data)):
            for col in range(len(field)):
                #因为第一行写了字段名称，所以要从第二行开始写入
                _ = sheet.cell(row=row+2, column=col + 1, value=u'%s' % data[row][col])
                #将生成的excel保存，这步是必不可少的
        newworkbook = new.save(file)
        #返回生成的excel
        return newworkbook

    def getYesterday(self,):
        # 获取昨天日期的字符串格式的函数
        #获取今天的日期
        today = datetime.date.today()
        #获取一天的日期格式数据
        oneday = datetime.timedelta(days=1)
        #昨天等于今天减去一天
        yesterday = today - oneday
        #获取昨天日期的格式化字符串
        yesterdaystr = yesterday.strftime('%Y-%m-%d')
        #返回昨天的字符串
        return yesterdaystr


    def create_email(self,email_from, email_to, email_Subject, email_text, annex_path, annex_name):
        return  self.m.create_email(email_from, email_to, email_Subject, email_text, annex_path, annex_name)

    def send_email(self,sender, password, receiver, msg):
        self.m.send_email(sender, password, receiver, msg)


if __name__ == "__main__":
    pass